import os
import re
import tkinter as tk
from tkinter import Toplevel, Label, Entry, Button, StringVar, Frame, Scrollbar, Text, CENTER, ttk, messagebox
from tkcalendar import DateEntry
from openpyxl import load_workbook
from datetime import datetime
from notas_app.utils.excel_utils import salvar_planilha
from notas_app.config import ARQUIVO_EXCEL, STATUS_COLETA

def mostrar_produtos(event, tree, janela_pai):
    selecionado = tree.selection()
    if not selecionado:
        return
    item = selecionado[0]
    valores = tree.item(item, "values")
    if len(valores) < 4:
        return
    produtos = valores[3]

    janela_produtos = Toplevel(janela_pai)
    janela_produtos.title(f"Produtos da Nota {valores[0]}")
    janela_produtos.geometry("500x400")
    janela_produtos.resizable(True, True)

    Label(janela_produtos, text="Produtos da Nota", font=("Segoe UI", 11, "bold")).pack(pady=5)

    scrollbar = Scrollbar(janela_produtos)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    text_area = Text(janela_produtos, wrap="word", yscrollcommand=scrollbar.set, font=("Segoe UI", 10))
    text_area.insert(tk.END, produtos)
    text_area.config(state="disabled")
    text_area.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
    scrollbar.config(command=text_area.yview)

def alterar_status(tree, wb, janela_pai, filtrar):
    selecionado = tree.selection()
    if not selecionado:
        return messagebox.showwarning("Aviso", "Selecione uma nota na lista para alterar o status.")
    item = selecionado[0]
    status_atual = tree.set(item, "status")

    status_popup = Toplevel(janela_pai)
    status_popup.title("Alterar Status")
    status_popup.geometry("250x120")
    status_popup.grab_set()

    Label(status_popup, text="Novo Status:", font=("Segoe UI", 10, "bold")).pack(pady=5)
    status_var_popup = StringVar(value=status_atual)
    status_combo_popup = ttk.Combobox(status_popup, textvariable=status_var_popup, values=STATUS_COLETA.split(","))
    status_combo_popup.pack(pady=5)

    def salvar():
        novo_status = status_var_popup.get()
        tree.set(item, "status", novo_status)
        nota_num = tree.item(item)["values"][0]
        for aba in wb.sheetnames:
            ws = wb[aba]
            for row in ws.iter_rows(min_row=2):
                if str(row[0].value) == str(nota_num):
                    row[9].value = novo_status
                    salvar_planilha(wb)
                    break
        status_popup.destroy()
        filtrar()
        tree.selection_set(item)  # re-seleciona item alterado

    Button(status_popup, text="Salvar", bg="green", fg="white", font=("Segoe UI", 10, "bold"), command=salvar).pack(pady=5)

def buscar_notas():
    try:
        wb = load_workbook(ARQUIVO_EXCEL)
    except Exception as e:
        messagebox.showerror("Erro", f"Planilha não encontrada ou em uso:\n{e}")
        return

    janela = Toplevel()
    janela.title("🔍 Consultar Notas Fiscais")
    janela.geometry("1000x650")

    Label(janela, text="🔍 Consultar Notas Fiscais", font=("Segoe UI", 13, "bold")).pack(pady=(10, 0))

    filtros_frame = Frame(janela)
    filtros_frame.pack(pady=10)
    pad_x = 25

    # Campos de filtro
    Label(filtros_frame, text="Data de Emissão:").grid(row=0, column=0, sticky="w", padx=pad_x)
    data_entry = DateEntry(filtros_frame, date_pattern="dd/mm/yyyy", width=16)
    data_entry.grid(row=1, column=0, padx=pad_x, pady=5)
    data_entry.delete(0, tk.END)
    data_entry._date = None

    Label(filtros_frame, text="Número da Nota:").grid(row=0, column=1, sticky="w", padx=pad_x)
    numero_entry = Entry(filtros_frame)
    numero_entry.grid(row=1, column=1, padx=pad_x, pady=5)

    Label(filtros_frame, text="Referência:").grid(row=0, column=2, sticky="w", padx=30)
    ref_entry = Entry(filtros_frame)
    ref_entry.grid(row=1, column=2, padx=pad_x, pady=5)

    Label(filtros_frame, text="Empresa Destinatária:").grid(row=2, column=0, sticky="w", padx=pad_x)
    empresa_entry = Entry(filtros_frame)
    empresa_entry.grid(row=3, column=0, padx=pad_x, pady=5)

    Label(filtros_frame, text="Valor Total da Nota:").grid(row=2, column=1, sticky="w", padx=pad_x)
    valor_entry = Entry(filtros_frame)
    valor_entry.grid(row=3, column=1, padx=pad_x, pady=5)

    Label(filtros_frame, text="Status Coleta:").grid(row=2, column=2, sticky="w", padx=pad_x)
    status_var = StringVar()
    status_combo = ttk.Combobox(filtros_frame, textvariable=status_var, values=STATUS_COLETA.split(","))
    status_combo.grid(row=3, column=2, padx=pad_x, pady=5)

    botoes_frame = Frame(janela)
    botoes_frame.pack(pady=5)

    tree_frame = Frame(janela)
    tree_frame.pack(padx=10, pady=10, fill="both", expand=True)
    scrollbar_y = Scrollbar(tree_frame, orient="vertical")
    scrollbar_y.pack(side=tk.RIGHT, fill="y")

    tree = ttk.Treeview(tree_frame,
                        columns=("nota", "data", "empresa", "produtos", "valor", "referencia", "status"),
                        show="headings", selectmode="browse", yscrollcommand=scrollbar_y.set)
    col_larguras = [90, 90, 160, 220, 80, 100, 90]
    for col, largura in zip(tree["columns"], col_larguras):
        tree.heading(col, text=col.title())
        tree.column(col, anchor=CENTER, width=largura)
    tree.pack(side="left", fill="both", expand=True)
    scrollbar_y.config(command=tree.yview)

    style = ttk.Style()
    style.configure("Treeview", rowheight=25, font=("Segoe UI", 10))
    style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
    style.map('Treeview', background=[('selected', '#9d9d9d')])

    tree.bind("<Double-1>", lambda e: mostrar_produtos(e, tree, janela))

    # Tags
    tree.tag_configure("evenrow", background="#f2f2f2")
    tree.tag_configure("oddrow", background="#ffffff")
    tree.tag_configure("evenrow_coletado", background="#b6f7b6")
    tree.tag_configure("oddrow_coletado", background="#d8fcd8")

    def filtrar():
        for item in tree.get_children():
            tree.delete(item)

        criterios = {
            "data": data_entry.get().strip(),
            "numero": numero_entry.get().strip(),
            "ref": ref_entry.get().strip().lower(),
            "empresa": empresa_entry.get().strip().lower(),
            "valor": valor_entry.get().strip(),
            "status": status_var.get().strip().lower()
        }

        zebra = False
        for aba in wb.sheetnames:
            ws = wb[aba]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                if isinstance(row[1], datetime):
                    data_formatada = row[1].strftime("%d/%m/%Y")
                elif isinstance(row[1], str) and re.match(r"\d{2}/\d{2}/\d{4}", row[1]):
                    data_formatada = row[1]
                else:
                    data_formatada = ""

                if not (
                    (criterios["data"] == data_formatada if criterios["data"] else True) and
                    (criterios["numero"] in str(row[0]) if criterios["numero"] else True) and
                    (criterios["ref"] in str(row[8]).lower() if criterios["ref"] else True) and
                    (criterios["empresa"] in str(row[2]).lower() if criterios["empresa"] else True) and
                    (criterios["valor"] == str(row[7]) if criterios["valor"] else True) and
                    (criterios["status"] == str(row[9]).lower() if criterios["status"] else True)
                ):
                    continue

                base_tag = "evenrow" if zebra else "oddrow"
                tag = base_tag + "_coletado" if str(row[9]).lower() == "coletado" else base_tag

                tree.insert(
                    "",
                    "end",
                    values=(row[0], data_formatada, row[2], row[3], row[7], row[8], row[9]),
                    tags=(tag,)
                )
                zebra = not zebra

    def limpar_filtros():
        data_entry.delete(0, tk.END)
        data_entry._date = None
        numero_entry.delete(0, tk.END)
        ref_entry.delete(0, tk.END)
        empresa_entry.delete(0, tk.END)
        valor_entry.delete(0, tk.END)
        status_var.set("")
        filtrar()

    def excluir_nota():
        item_selecionado = tree.selection()
        if not item_selecionado:
            messagebox.showwarning("Aviso", "Selecione uma nota para excluir.")
            return

        confirmacao = messagebox.askyesno("Confirmar exclusão", "Tem certeza que deseja excluir esta nota?")
        if not confirmacao:
            return

        numero_nota = tree.item(item_selecionado, "values")[0]
        try:
            wb_local = load_workbook(ARQUIVO_EXCEL)
            for aba in wb_local.sheetnames:
                ws = wb_local[aba]
                for row in ws.iter_rows(min_row=2):
                    if str(row[0].value) == str(numero_nota):
                        ws.delete_rows(row[0].row, 1)
                        wb_local.save(ARQUIVO_EXCEL)
                        break
            tree.delete(item_selecionado)
            messagebox.showinfo("Sucesso", f"Nota {numero_nota} excluída com sucesso.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao excluir nota: {e}")

    # Botões
    btn_limpar = Button(botoes_frame, text="🧹 Limpar Filtros", cursor="hand2",
                        bg="darkgoldenrod", fg="white", font=("Segoe UI", 10),
                        command=limpar_filtros)
    btn_limpar.pack(side="left", padx=10)

    btn_status = Button(botoes_frame, text="🔄 Alterar Status", cursor="hand2",
                        bg="blue", fg="white", font=("Segoe UI", 10),
                        command=lambda: alterar_status(tree, wb, janela, filtrar))
    btn_status.pack(side="left", padx=5)

    btn_excluir = Button(botoes_frame, text="🗑 Excluir Nota", cursor="hand2",
                         bg="red", fg="white", font=("Segoe UI", 10),
                         command=excluir_nota)
    btn_excluir.pack(side="left", padx=10)

    # Bind filtros
    for widget in [data_entry, numero_entry, ref_entry, empresa_entry, valor_entry]:
        widget.bind("<KeyRelease>", lambda e: filtrar())
    status_combo.bind("<<ComboboxSelected>>", lambda e: filtrar())

    filtrar()
