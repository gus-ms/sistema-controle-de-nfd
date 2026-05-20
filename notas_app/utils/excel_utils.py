from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.worksheet.datavalidation import DataValidation

from notas_app.config import ARQUIVO_EXCEL, CABECALHOS, COL_WIDTHS, STATUS_COLETA

# Bordas
def definir_bordas():
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    thick = Border(left=Side(style='medium'), right=Side(style='medium'),
                   top=Side(style='medium'), bottom=Side(style='medium'))
    return thin, thick

def criar_aba_mes(wb, mes_ano):
    if mes_ano not in wb.sheetnames:
        ws = wb.create_sheet(title=mes_ano)
        ws.append(CABECALHOS)
        bold_font = Font(bold=True)
        thin_border, thick_border = definir_bordas()
        for col in range(1, len(CABECALHOS)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = bold_font
            cell.border = thick_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for i, width in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[chr(64+i)].width = width
        ws.row_dimensions[1].height = 25
        # Validação para Status
        dv_status = DataValidation(type="list", formula1=f'"{STATUS_COLETA}"')
        ws.add_data_validation(dv_status)
        dv_status.add('J2:J1048576')

def salvar_planilha(wb):
    wb.save(ARQUIVO_EXCEL)
