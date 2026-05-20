import os, re, shutil
import xml.etree.ElementTree as ET
from datetime import datetime
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from notas_app.config import PASTA_PROCESSADOS, ARQUIVO_EXCEL
from notas_app.utils.excel_utils import criar_aba_mes, salvar_planilha
from notas_app.config import ARQUIVO_EXCEL, STATUS_COLETA, PASTA_XML, PASTA_PROCESSADOS
from notas_app.utils.files_utils import renomear_arquivo

def extrair_dados_xml(caminho):
    tree = ET.parse(caminho)
    root = tree.getroot()
    ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

    numero_nota = root.find('.//nfe:ide/nfe:nNF', ns).text
    data_emissao_str = root.find('.//nfe:ide/nfe:dhEmi', ns).text[:10]
    data_emissao = datetime.strptime(data_emissao_str, '%Y-%m-%d').date()
    nome_dest = root.find('.//nfe:dest/nfe:xNome', ns).text
    valor_total_nota = float(root.find('.//nfe:total/nfe:ICMSTot/nfe:vNF', ns).text)

    referencia = "NFE NAO ENCONTRADA"
    infCpl = root.find('.//nfe:infAdic/nfe:infCpl', ns)
    if infCpl is not None and infCpl.text:
        match = re.search(r'\bNF[E]?\s*(\d+)', infCpl.text.upper())
        if match:
            referencia = f"NFE {match.group(1)}"

    produtos = []
    for det in root.findall('.//nfe:det', ns):
        prod = det.find('.//nfe:prod', ns)
        if prod is not None:
            nome = prod.find('nfe:xProd', ns).text
            qtd = prod.find('nfe:qCom', ns).text
            un = prod.find('nfe:uCom', ns).text
            produtos.append(f"{nome} ({qtd} {un})")
    produtos_str = '\n'.join(produtos)

    quantidade_total = sum(
        float(det.find('.//nfe:prod/nfe:qCom', ns).text)
        for det in root.findall('.//nfe:det', ns)
    )
    valor_unitario_medio = valor_total_nota / quantidade_total if quantidade_total else 0

    return {
        "numero_nota": numero_nota,
        "data_emissao": data_emissao,
        "nome_dest": nome_dest,
        "valor_total_nota": valor_total_nota,
        "referencia": referencia,
        "produtos_str": produtos_str,
        "quantidade_total": quantidade_total,
        "valor_unitario_medio": round(valor_unitario_medio, 2),
    }

def inserir_nota_planilha(wb, dados):
    mes_ano = dados["data_emissao"].strftime('%b-%Y').lower()
    criar_aba_mes(wb, mes_ano)
    ws = wb[mes_ano]

    existe = any(
        str(ws.cell(row=row, column=1).value) == str(dados["numero_nota"])
        for row in range(2, ws.max_row + 1)
    )
    if existe:
        return False

    linha_vazia = None
    for row in range(2, ws.max_row + 10):
        if all(ws.cell(row=row, column=col).value is None for col in range(1, 12)):
            linha_vazia = row
            break
    if linha_vazia is None:
        linha_vazia = ws.max_row + 1

    linha = [
        dados["numero_nota"],
        dados["data_emissao"],
        dados["nome_dest"],
        dados["produtos_str"],
        dados["quantidade_total"],
        dados["valor_unitario_medio"],
        dados["valor_total_nota"],
        dados["valor_total_nota"],
        dados["referencia"],
        "Em coleta",
    ]

    for col, val in enumerate(linha, 1):
        ws.cell(row=linha_vazia, column=col).value = val

    ws.cell(row=linha_vazia, column=2).number_format = 'DD/MM/YYYY'
    ws.cell(row=linha_vazia, column=4).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[linha_vazia].height = 25

    return True

def processar_xmls(lista_arquivos):
    if not lista_arquivos:
        return

    os.makedirs(PASTA_PROCESSADOS, exist_ok=True)

    try:
        wb = load_workbook(ARQUIVO_EXCEL)
    except FileNotFoundError:
        wb = Workbook()

    notas_processadas = []
    notas_duplicadas = []
    erros_importacao = []

    for caminho in lista_arquivos:
        try:
            dados = extrair_dados_xml(caminho)
            inseriu = inserir_nota_planilha(wb, dados)
            if not inseriu:
                notas_duplicadas.append(dados["numero_nota"])
                continue

            novo_caminho = renomear_arquivo(caminho, dados["numero_nota"], dados["nome_dest"])
            shutil.move(caminho, novo_caminho)
            notas_processadas.append(os.path.basename(novo_caminho))

        except Exception as e:
            erros_importacao.append(f"{os.path.basename(caminho)}: {e}")

    salvar_planilha(wb)

    mensagens = []
    if notas_processadas:
        mensagens.append(f"Notas importadas:\n" + "\n".join(notas_processadas))
    if notas_duplicadas:
        mensagens.append(f"Notas duplicadas (ignoradas):\n" + "\n".join(notas_duplicadas))
    if erros_importacao:
        mensagens.append(f"Erros na importação:\n" + "\n".join(erros_importacao))

    mensagem_final = "\n\n".join(mensagens) if mensagens else "Nenhuma nota foi processada."
    messagebox.showinfo("Importação Finalizada", mensagem_final)
